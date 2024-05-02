from openpyxl import Workbook
import pandas as pd

def createReport():
    csv = 'vacancies.csv'
    df=pd.read_csv(csv, header=None)
    df.columns=['name','salary_from','salary_to','salary_currency','area_name','date_time']
    df['salary'] = (df['salary_from'] + df['salary_to']) / 2
    year_df=createYearDataFrame(df)
    city_salary_df,city_percent_df=createCityDf(df)
    wb=createWb(year_df,city_salary_df,city_percent_df)
    wb.save('student_works/test.xlsx')
def createYearDataFrame(df):
    df['year']=df['date_time'].str.split('-').str[0]
    average_salary = df.groupby('year')['salary'].mean().round().reset_index()
    vacancies_per_year = df.groupby('year')['name'].count().reset_index()
    average_salary['count'] = vacancies_per_year['name']
    return average_salary
def createCityDf(df):
    city_counts=df['area_name'].value_counts()
    city_percent = (city_counts / len(df)*100).round(2)
    cities_over_one_percent = city_percent[city_percent > 1].index
    average_salary=df[df['area_name'].isin(cities_over_one_percent)].groupby('area_name')['salary'].mean().sort_values(ascending=False).round()
    print(average_salary[:10].reset_index(),city_percent[:10].reset_index())
    return average_salary[:10].reset_index(),city_percent[:10].reset_index()
def createWb(year_df,city_salary_df,city_percent_df):
    wb = Workbook()
    wb.create_sheet(title='Статистика по годам', index=0)
    wb.create_sheet(title='Статистика по городам', index=1)
    sheetnames = wb.sheetnames 
    first_sheet = wb[sheetnames[0]]
    second_sheet = wb[sheetnames[1]]
    wb.remove(wb[sheetnames[2]])
    first_sheet[f"A{1}"], first_sheet[f"B{1}"], first_sheet[f"C{1}"] = 'Год', 'Средняя зарплата', 'Количество вакансий'
    second_sheet[f"A{1}"], second_sheet[f"B{1}"], second_sheet[f"D{1}"], second_sheet[
        f"E{1}"] = 'Город', 'Уровень зарплат', 'Город', 'Доля вакансий, %'
    createYearPage(year_df,first_sheet)
    createCityPage(city_salary_df,city_percent_df,second_sheet)
    return wb

def createYearPage(df,page):
    for i in range(len(df)):
        year, salary, count = int(df['year'][i]),df['salary'][i],df['count'][i] 
        page[f"A{i+2}"] = year
        page[f"B{i+2}"] = salary
        page[f"C{i+2}"] = count
def createCityPage(city_salary_df,city_percent_df,page):
    for i in range(len(city_salary_df)):
        town, salary,city, percent = city_salary_df['area_name'][i],city_salary_df['salary'][i],city_percent_df['area_name'][i],city_percent_df['count'][i] 
        page[f"A{i+2}"] = town
        page[f"B{i+2}"] = salary
        page[f"D{i+2}"] = city
        page[f"E{i+2}"] = percent

createReport()
#print(vacancies_per_year)
from openpyxl import Workbook
import pandas

def createReport():
    csv = 'vacancies.csv'
    vacancies_df=pandas.read_csv(csv, header=None)
    vacancies_df.columns=['name','salary_from','salary_to','salary_currency','area_name','date_time']
    vacancies_df['salary'] = (vacancies_df['salary_from'] + vacancies_df['salary_to']) / 2
    year_df=createYearDataFrame(vacancies_df)
    city_salary_df,city_percent_df=createCityDataFrame(vacancies_df)
    wb=createWorkbook(year_df,city_salary_df,city_percent_df)
    wb.save('student_works/report.xlsx')

def createYearDataFrame(vacancies_df):
    vacancies_df['date_time'] = pandas.to_datetime(vacancies_df['date_time'], errors='coerce', utc=True)
    vacancies_df['year'] = vacancies_df['date_time'].dt.year
    yearly_data = vacancies_df.groupby('year').agg({'salary': 'mean', 'name': 'count'}).reset_index()
    yearly_data['salary'] = yearly_data['salary'].round()
    yearly_data.rename(columns={'name': 'count'}, inplace=True)
    return yearly_data

def createCityDataFrame(vacancies_df):
    city_data = vacancies_df['area_name'].value_counts(normalize=True).mul(100).round(2)
    cities_over_one_percent = city_data[city_data > 1].index
    average_salary = vacancies_df[vacancies_df['area_name'].isin(cities_over_one_percent)].groupby('area_name')['salary'].mean().sort_values(ascending=False).round()
    top_cities_salary = average_salary[:10].reset_index()
    top_cities_percent = city_data[:10].reset_index().rename(columns={'area_name': 'percent'})
    return top_cities_salary, top_cities_percent

def createWorkbook(year_df, city_salary_df, city_percent_df):
    wb = Workbook()
    year_sheet = wb.create_sheet(title='Статистика по годам', index=0)
    city_sheet = wb.create_sheet(title='Статистика по городам', index=1)
    wb.remove(wb['Sheet'])
    createYearPage(year_df, year_sheet)
    createCityPage(city_salary_df, city_percent_df, city_sheet)
    return wb

def createYearPage(year_df, sheet):
    sheet.append(['Год', 'Средняя зарплата', 'Количество вакансий'])
    for row in year_df.itertuples(index=False):
        sheet.append(row)

def createCityPage(city_salary_df, city_percent_df, sheet):
    sheet.append(['Город', 'Уровень зарплат', None, 'Город', 'Доля вакансий, %'])
    for salary_row, percent_row in zip(city_salary_df.itertuples(index=False), city_percent_df.itertuples(index=False)):
        sheet.append(salary_row + (None,) + percent_row)

createReport()